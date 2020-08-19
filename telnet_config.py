import openpyxl
import telnetlib
import subprocess
import time
#
PATH = '/Users/alawn/Desktop/test.xlsx'
Workbook = openpyxl.load_workbook(PATH)
Sheet = Workbook['Sheet1']
row = Sheet.max_row
username = input('Username:')
Password = input('Password:')
enable = input('enable:')
def get_device_info():
    for n in range (1,row+1):
        ip = Sheet.cell(row=n,column=1).value
        port = Sheet.cell(row=n,column=2).value
        config = str(Sheet.cell(row=n,column=3).value).split('\n')
        yield ip,port,config
device_info = get_device_info()

def config_device(host,port,commonds):
    tn = telnetlib.Telnet(host,port=port)
    tn.set_debuglevel(5)
    tn.write(b'\r\n')
    tn.read_until(b'Username:')
    time.sleep(1)
    tn.write(username.encode()+b'\r\n')
    time.sleep(1)
    tn.read_until(b'Password:')
    time.sleep(1)
    tn.write(Password.encode()+b'\r\n')
    time.sleep(1)
    tn.read_until(b'>')
    time.sleep(1)
    tn.write(b'enable\r\n')
    time.sleep(1)
    tn.read_until(b'Password:')
    time.sleep(1)
    tn.write(enable.encode()+b'\r\n')
    time.sleep(1)
    tn.read_until(b'#')
    time.sleep(1)
    tn.write(b'conf t\r\n')
    time.sleep(1)
    tn.read_until(b')#')
    for commond in commonds:
        tn.write((commond.encode())+b'\n')
    time.sleep(1)
    tn.read_until(b'#')
    time.sleep(1)
    tn.write(b'end\r\n')
    time.sleep(1)
    tn.read_until(b'#')
    time.sleep(1)
    tn.write(b'exit\r\n')
    time.sleep(1)
    tn.close()

if __name__ == '__main__':
    for n in device_info:
        host = n[0]
        port = n[1]
        commonds = n[2]
        config_device(host,port,commonds)



