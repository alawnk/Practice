import os
from openpyxl import load_workbook
from collections import defaultdict
from IPy import IP
PATH  = '/Users/alawn/Desktop/生态园IP规划表.xlsx'
WORKBOOK = load_workbook(PATH)
IP_PLANNING = WORKBOOK['有线IP规划表']
CONNECTED_RELATION = WORKBOOK['连接关系表']
IP_MAX_ROWS = IP_PLANNING.max_row
CR_MAX_ROWS = CONNECTED_RELATION.max_row

master_template = '''
interface vlan $VLAN_NO
description  $DESC
ip address $M_IP_ADDR $MASK
standby version 2
standby $VLAN_NO ip $HSRP_IP
standby $VLAN_NO priority 120
standby $VLAN_NO preempt
'''


backup_template = '''
interface vlan $VLAN_NO
description  $DESC
ip address $B_IP_ADDR $MASK
standby version 2
standby $VLAN_NO ip $HSRP_IP
standby $VLAN_NO priority 100
standby $VLAN_NO preempt
'''
class IP_PLAN:
    def __init__(self,vlan,network,function,floor,mip,bip,hip,mask):
        self.vlan = vlan
        self.network = network
        self.function = function
        self.floor = floor
        self.master_ip = mip
        self.backup_ip = bip
        self.hsrp_ip = hip
        self.mask = mask

def network_values():
    Floor = []

    for n in range (2,IP_MAX_ROWS+1):

        vlan = IP_PLANNING.cell(row=n,column=1).value
        network = IP_PLANNING.cell(row=n,column=2).value
        function = IP_PLANNING.cell(row=n,column=3).value
        floor = IP_PLANNING.cell(row=n,column=5).value
        hip = IP(int(IP(network).strDec())+1).strNormal(0)
        mip = IP(int(IP(network).strDec())+2).strNormal(0)
        bip = IP(int(IP(network).strDec())+3).strNormal(0)
        mask = IP(network).netmask()
        if floor != None:
            floor_new = floor.strip()

            res = IP_PLAN(vlan,network,function,floor_new,hip,mip,bip,mask)

            if res.__dict__['floor'] != None:
                Floor.append(res.__dict__)
            else:
                pass
    return Floor

def new_kv():
    new_dict = {}
    for k in network_values():
        print(k.keys())
new_kv()
        # new_dict.setdefault(k['floor'],[]).append([k['vlan'],k['network'],k['function'],k['floor'],k['master_ip'],k['backup_ip'],k['hsrp_ip'],k['mask']])
    # return new_dict

# for value in new_kv():
#     print(new_kv()[value])





