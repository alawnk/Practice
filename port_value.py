import os
import openpyxl
PATH = '/Users/alawn/Desktop/result-1/'
for root,dir,files in os.walk(PATH):
    for file in files:
        # print(file)
        workbook  = openpyxl.load_workbook(PATH+file)
        sheets = workbook.sheetnames
        for i in range (len(sheets)):
            sheet = workbook[sheets[i]]
            port=sheet['A2'].value
            # print(port)
            row = sheet.max_row
            col = sheet.max_column
            input = []
            output =[]
            for n in range (2,row+1):
                input_value = sheet.cell(row=n,column = 3).value
                input.append(input_value)
                output_value = sheet.cell(row=n,column=4).value
                output.append(output_value)
def write_excel():
    port_excel = openpyxl.load_workbook('/Users/alawn/Desktop/Port.xlsx')
    work_port = port_excel['Sheet1']
    v = 0
    work_port.cell(row=n,column=1,value=file)
    work_port.cell(row=n,column=2,value=port)
    work_port.cell(row=n,column=3,value=max(input))
    work_port.cell(row=n,column=4,value=max(output))
    v += 1
    port_excel.save('Port.xlsx')
write_excel()





