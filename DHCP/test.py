import openpyxl
import xlrd

workbook = xlrd.open_workbook('D:/GITHUB/DHCP/change.xlsx')
sheet = workbook.sheet_by_index(0)
nrows = sheet.nrows
wb = openpyxl.load_workbook('D:/GITHUB/DHCP/change.xlsx')
ws = wb['Sheet1']
for i in range (0,nrows):
    c1 = sheet.row(i)[1].value
    c2 = sheet.row(i)[2].value
    c3 = sheet.row(i)[3].value
    c4 = sheet.row(i)[5].value
    c5 = sheet.row(i)[6].value
    c6 = sheet.row(i)[7].value
    c7 = sheet.row(i)[8].value
    times = xlrd.xldate.xldate_as_datetime(c7,0)
    c8 = sheet.row(i)[9].value
    pro_info = ('%s-%s-%s'%(c1,c2,c3))
    project_info=('''[变更项目]:%s
[变更人员]:%s
[变更原因]:%s
[变更描述]:%s
[变更时间]:%s
[变更影响]:%s'''%(pro_info,c4,c5,c6,times,c8))
    C = [project_info]
    for c in C:
        ws.cell(i+1,ws.max_column).value = c
wb.save('D:/GITHUB/DHCP/change.xlsx')














































































# pro_info = '有线—配置修改-配置'
# time =  '2020年1月1日 19：00'
# res = '沃尔沃二为'
# cas = 'werwe'
# print('''[变更项目]:%s
# [变更时间]:%s
# [变更原因]:%s
# [变更影响]:%s'''%(pro_info,time,res,cas))\n