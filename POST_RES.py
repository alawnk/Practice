import requests
import openpyxl
import queue
import threading
url = " "
path = 'C:/Users/alawnwang/Desktop/Device_Info.xlsx'
txt_path = 'D:/EB_POST_RES_QUEUE/'
wb = openpyxl.load_workbook(path)
ws = wb['鹅埠接入']
row = ws.max_row

device_ids = []
device_names = []

for n in range(2, row + 1):
    device_id = ws.cell(row=n, column=1).value
    device_name = ws.cell(row=n, column=4).value
    device_ids.append(device_id)
    device_names.append(device_name)

myq = queue.Queue()
for id,name in zip(device_ids,device_names):
    myq.put((id, name))

def get_post_res():
    while True:
        try:
            info = myq.get(block=False)
            payload =  "{\r\n    \"method\": \"snmp.get_snmp_data\",\r\n    \"systemid\":\" \",\r\n    \"params\": {\r\n        \"start_date\":\"2020-12-28 00:00:00\",\r\n        \"end_date\":\"2021-01-01 23:59:59\",\r\n        \"devid\":%s,\r\n        \"attr\":[\"in_traffic\"]\r\n    }\r\n}" %info[0]
            headers = {'Content-Type': 'text/plain'}
            res = open(txt_path+info[1],'w')
            response = requests.request("POST", url, headers=headers, data = payload)
            res.write(response.text)
        except queue.Empty:
            break

thread_arr = []

num_thread = 20

for i in range (num_thread):
        t = threading.Thread(target = get_post_res,kwargs={})
        t.setDaemon(True)
        t.start()
        thread_arr.append(t)

for i in range(num_thread):
        thread_arr[i].join()
